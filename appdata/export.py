'''
Diese Script zipped den Upload-Ordner für den Drucker/Sigloch
Nach dem Zippen bitte das Zip File kopieren
'''

import os
import shutil
import sqlite3
import qrcode
import openpyxl
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime
from time import strftime
import re

db1 = sqlite3.connect("db/db.db")

sql1 = """ select costumer_name, name, streetname, plz, city, email, phone, country, session_id, file_data, opt3 FROM (select * from data_costumer order by country collate nocase, costumer_name collate nocase) group by session_id order by country collate nocase, costumer_name collate nocase"""
sql2 = """ SELECT company, salutation, firstname, lastname, street, number, zipcode, city, country, qr, url, division, address_supplement from (select * FROM data_retailer WHERE session_id = ? order by country collate nocase, company collate nocase) group by company, salutation, firstname, lastname, street, number, zipcode, city, country """
sql3 = """ SELECT COUNT(*) FROM (SELECT distinct company, salutation, firstname, lastname, street, number, zipcode, city, country FROM data_retailer) """


def get_retailer(session_id):
    cy = db1.cursor()
    cy.execute(sql2, (session_id, ))
    return cy.fetchall()


def getnow():
    now = datetime.now()
    return now.strftime("%Y_%m_%d-%H_%M")


def get_salutation(salutation, firstname, lastname, division):
    women_regex = "(Frau|geehrte +|Liebe +|Damen)"
    men_regex = "(Herr|geehrter +|Lieber +|Mann)"
    division_regex = "(Abteilung|abteilung|Abt\.|abt\.)"

    salutation_string = get_value(salutation)
    firstname_string = get_value(firstname)
    lastname_string = get_value(lastname)
    division_string = get_value(division)

    if salutation_string and firstname_string and lastname_string:
        if matches(salutation_string, women_regex) and not matches(salutation_string, men_regex):
            return f"Liebe {firstname_string} {lastname_string}"
        elif not matches(salutation_string, women_regex) and matches(salutation_string, men_regex):
            return f"Lieber {firstname_string} {lastname_string}"
        return f"Liebe(r) {firstname_string} {lastname_string}"

    if salutation_string and firstname_string and not lastname_string:
        if matches(salutation_string, women_regex) and not matches(salutation_string, men_regex):
            return f"Liebe {firstname_string}"
        elif not matches(salutation_string, women_regex) and matches(salutation_string, men_regex):
            return f"Lieber {firstname_string}"
        return f"Liebe(r) {firstname_string}"

    if salutation_string and not firstname_string and lastname_string:
        if matches(salutation_string, women_regex) and not matches(salutation_string, men_regex):
            return f"Liebe Frau {lastname_string}"
        elif not matches(salutation_string, women_regex) and matches(salutation_string, men_regex):
            return f"Lieber Herr {lastname_string}"
        return f"Liebe(r) Frau/Herr {lastname_string}"

    if not salutation_string and firstname_string and lastname_string:
        if matches(firstname_string, women_regex) and not matches(firstname_string, men_regex):
            return f"Liebe {firstname_string} {lastname_string}"
        if not matches(firstname_string, women_regex) and matches(firstname_string, men_regex):
            return f"Lieber {firstname_string} {lastname_string}"
        return f"Liebe(r) {firstname_string} {lastname_string}"

    if not salutation_string and firstname_string and not lastname_string:
        if matches(firstname_string, women_regex) and not matches(firstname_string, men_regex):
            return f"Liebe {firstname_string}"
        if not matches(firstname_string, women_regex) and matches(firstname_string, men_regex):
            return f"Lieber {firstname_string}"
        return f"Liebe(r) {firstname_string}"

    if not salutation_string and not firstname_string and lastname_string:
        if matches(lastname_string, women_regex) and not matches(lastname_string, men_regex):
            return f"Liebe {lastname_string}"
        if not matches(lastname_string, women_regex) and matches(lastname_string, men_regex):
            return f"Lieber {lastname_string}"
        return f"Liebe(r) Frau/Herr {lastname_string}"

    if not salutation_string and not firstname_string and not lastname_string and division_string:
        if matches(division_string, division_regex):
            return f"Liebe {division_string}"
        return f"Liebe Abteilung {division_string}"

    return "Liebe Teamplayer*innen"


def matches(value, regex):
    return re.search(regex, value)


def get_value(string):
    if string:
        return string.strip()
    return ""


cx = db1.cursor()
cx.execute(sql1)
rx = cx.fetchall()

cz = db1.cursor()
cz.execute(sql3)
rz = cz.fetchone()
print(rz[0])

wb = Workbook()
ws = wb.active
ws.title = "HAKRO Fachhändler"

now = getnow()
source_file = "static/export/export.xlsx"
target_file = f"static/export/export_{now}.xlsx"
wbx = load_workbook(filename=source_file)
ws = wbx['Tabelle1']
s = wbx.active

START_ROW = 2
MAX_ROW = int(rz[0]) + 1

ix = START_ROW
for i in rx:
    print(i)
    retailers = get_retailer(i[8])
    for r in retailers:
        s.cell(row=ix, column=1).value = i[8]   # session_id
        s.cell(row=ix, column=2).value = i[0]   # costumer_name
        s.cell(row=ix, column=3).value = i[1]   # name
        s.cell(row=ix, column=4).value = i[2]   # streetname
        s.cell(row=ix, column=5).value = i[3]   # plz
        s.cell(row=ix, column=6).value = i[4]   # city
        s.cell(row=ix, column=7).value = i[5]   # email
        s.cell(row=ix, column=8).value = i[6]   # phone
        s.cell(row=ix, column=9).value = i[9]   # file_data
        s.cell(row=ix, column=10).value = f'{i[8]}.pdf'  # file_logo
        s.cell(row=ix, column=11).value = i[7]  # country
        s.cell(row=ix, column=12).value = i[10]  # opt3
        # Kunden
        s.cell(row=ix, column=13).value = r[0]  # company
        s.cell(row=ix, column=14).value = r[11]  # division
        s.cell(row=ix, column=15).value = get_salutation(
            salutation=r[1], firstname=r[2], lastname=r[3], division=r[11])  # salutation
        s.cell(row=ix, column=16).value = r[2]  # firstname
        s.cell(row=ix, column=17).value = r[3]  # lastname
        s.cell(row=ix, column=18).value = r[4]  # street
        s.cell(row=ix, column=19).value = r[5]  # number
        s.cell(row=ix, column=20).value = r[12]  # address supplement
        s.cell(row=ix, column=21).value = r[6]  # zipcode
        s.cell(row=ix, column=22).value = r[7]  # city
        s.cell(row=ix, column=23).value = r[8]  # country
        s.cell(row=ix, column=24).value = r[9]  # qr
        s.cell(row=ix, column=25).value = r[10]  # url
        s.cell(row=ix, column=26).value = r[1]

        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=15,
            border=1,
        )
        qr.add_data(f'{r[10]}')
        qr.make(fit=True)
        qr_filename = os.path.join('static', 'export', 'qr', f'{r[9]}.png')
        qr_code = qr.make_image(fill_color="black", back_color="white")
        qr_code.save(qr_filename)
        ix += 1

print("Saving workbook")
wbx.save(filename=target_file)

p = os.path.join('static', 'export')

print("Creating assets.zip")
output_1 = p + "/" + f'asstets_{now}'
shutil.make_archive(output_1, 'zip', os.path.join('static', 'upload'))

print("Creating db.zip")
output_2 = p + "/" + f'db_{now}'
shutil.make_archive(output_2, 'zip', os.path.join('db'))

print("Creating qr.zip")
output_3 = p + "/" + f'qr_{now}'
shutil.make_archive(output_3, 'zip', os.path.join('static', 'export', 'qr'))

# time.sleep(1)

# print("Creating all.zip")
# output_4 = p + "/" + f'all_{now}'
# shutil.make_archive(output_4, 'zip', os.path.join('static', 'export'))

print("Export finished")
